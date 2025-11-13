"""
Excel file loader for ontology annotations and metadata.

This module provides functionality to load Excel files that contain
detailed annotations and metadata for ontology entities.
"""

from pathlib import Path
from typing import Any, Dict, List, Optional

import structlog

logger = structlog.get_logger(__name__)


class ExcelLoader:
    """
    Loader for Excel files containing ontology annotations.
    
    This class can parse Excel files and extract annotations, descriptions,
    and metadata for ontology entities.
    """
    
    def __init__(self):
        self.logger = logger.bind(component="ExcelLoader")
    
    def load_excel_file(self, file_path: str) -> Dict[str, Any]:
        """
        Load an Excel file and extract annotations.
        
        Args:
            file_path: Path to the Excel file
            
        Returns:
            Dictionary with annotations and metadata
        """
        file_path = Path(file_path)
        if not file_path.exists():
            raise FileNotFoundError(f"Excel file not found: {file_path}")
        
        self.logger.info("Loading Excel file", file_path=str(file_path))
        
        try:
            # Try to import openpyxl or pandas
            try:
                import pandas as pd
                use_pandas = True
            except ImportError:
                try:
                    import openpyxl
                    use_pandas = False
                except ImportError:
                    self.logger.warning("Neither pandas nor openpyxl available. Excel loading disabled.")
                    return {
                        "annotations": {},
                        "metadata": {},
                        "error": "Excel libraries not installed. Install pandas or openpyxl."
                    }
            
            if use_pandas:
                return self._load_with_pandas(file_path)
            else:
                return self._load_with_openpyxl(file_path)
                
        except Exception as e:
            self.logger.error("Failed to load Excel file", error=str(e))
            return {
                "annotations": {},
                "metadata": {},
                "error": str(e)
            }
    
    def _load_with_pandas(self, file_path: Path) -> Dict[str, Any]:
        """Load Excel file using pandas."""
        import pandas as pd
        
        # Read all sheets
        excel_file = pd.ExcelFile(file_path)
        annotations = {}
        metadata = {}
        
        for sheet_name in excel_file.sheet_names:
            df = pd.read_excel(excel_file, sheet_name=sheet_name)
            
            # Convert DataFrame to dictionary
            sheet_data = []
            for _, row in df.iterrows():
                row_dict = {}
                for col in df.columns:
                    value = row[col]
                    # Handle NaN values
                    if pd.notna(value):
                        row_dict[str(col)] = str(value)
                if row_dict:
                    sheet_data.append(row_dict)
            
            # Try to identify entity name column
            entity_column = None
            for col in df.columns:
                col_lower = str(col).lower()
                if any(keyword in col_lower for keyword in ['entity', 'name', 'class', 'individual', 'concept']):
                    entity_column = col
                    break
            
            # Organize by entity if possible
            if entity_column:
                for row_dict in sheet_data:
                    entity_name = row_dict.get(str(entity_column), "Unknown")
                    if entity_name not in annotations:
                        annotations[entity_name] = {}
                    annotations[entity_name].update(row_dict)
            else:
                # Store as list if no entity column found
                metadata[sheet_name] = sheet_data
        
        self.logger.info("Excel file loaded", 
                        sheets=len(excel_file.sheet_names),
                        annotations_count=len(annotations))
        
        return {
            "annotations": annotations,
            "metadata": metadata,
            "sheets": excel_file.sheet_names,
        }
    
    def _load_with_openpyxl(self, file_path: Path) -> Dict[str, Any]:
        """Load Excel file using openpyxl."""
        from openpyxl import load_workbook
        
        workbook = load_workbook(file_path, data_only=True)
        annotations = {}
        metadata = {}
        
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            
            # Get headers from first row
            headers = []
            for cell in sheet[1]:
                headers.append(str(cell.value) if cell.value else "")
            
            # Read data rows
            sheet_data = []
            for row in sheet.iter_rows(min_row=2, values_only=True):
                row_dict = {}
                for i, value in enumerate(row):
                    if i < len(headers) and value is not None:
                        row_dict[headers[i]] = str(value)
                if row_dict:
                    sheet_data.append(row_dict)
            
            # Try to identify entity name column
            entity_column = None
            for i, header in enumerate(headers):
                header_lower = header.lower()
                if any(keyword in header_lower for keyword in ['entity', 'name', 'class', 'individual', 'concept']):
                    entity_column = i
                    break
            
            # Organize by entity if possible
            if entity_column is not None:
                for row_dict in sheet_data:
                    entity_name = list(row_dict.values())[entity_column] if entity_column < len(list(row_dict.values())) else "Unknown"
                    if entity_name not in annotations:
                        annotations[entity_name] = {}
                    annotations[entity_name].update(row_dict)
            else:
                metadata[sheet_name] = sheet_data
        
        self.logger.info("Excel file loaded", 
                        sheets=len(workbook.sheetnames),
                        annotations_count=len(annotations))
        
        return {
            "annotations": annotations,
            "metadata": metadata,
            "sheets": workbook.sheetnames,
        }
    
    def apply_annotations_to_entities(self, annotations: Dict[str, Any], 
                                      entities: Dict[str, Any]) -> Dict[str, Any]:
        """
        Apply Excel annotations to entities.
        
        Args:
            annotations: Dictionary of annotations from Excel
            entities: Dictionary of entities to annotate
            
        Returns:
            Updated entities dictionary with annotations
        """
        updated_entities = entities.copy()
        
        for entity_id, entity_data in updated_entities.items():
            entity_name = entity_data.get("name", "")
            
            # Try exact match first
            if entity_name in annotations:
                entity_data["properties"] = entity_data.get("properties", {})
                entity_data["properties"].update({
                    "excel_annotations": annotations[entity_name]
                })
            else:
                # Try case-insensitive match
                for ann_name, ann_data in annotations.items():
                    if ann_name.lower() == entity_name.lower():
                        entity_data["properties"] = entity_data.get("properties", {})
                        entity_data["properties"].update({
                            "excel_annotations": ann_data
                        })
                        break
        
        return updated_entities

